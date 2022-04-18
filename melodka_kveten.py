import re
from openpyxl import Workbook

import urllib.request, urllib.parse, urllib.error
import requests
from bs4 import BeautifulSoup

url = "https://melodka.cz/program/default/2022-5"
page = requests.get(url)
soup = BeautifulSoup(page.content, "html.parser")

html = urllib.request.urlopen(url).read()
print(type(html.decode()))
wb = Workbook()
ws1 = wb.active
ws1.title = "Kveten"

a = soup.find_all("div", class_="datum")
for i,j in enumerate(a):
    print(j.text)
    k = j.text.replace(" ","")
    k = k.rstrip("2022")
    ws1.cell(row=i+1, column=1).value = k

b = soup.find_all("div", class_="den")
for i,j in enumerate(b):
    print(j.text)
    ws1.cell(row=i+1, column=1).value += " "
    ws1.cell(row=i+1, column=1).value += j.text

#print(soup.text) vytiskne stránku textově ne html
c = re.findall("<div class=\"nazev[^>]*><a href=\"/program/akce/[^>]*>([^<]*)", html.decode())
for i,j in enumerate(c):
    print(j)
    ws1.cell(row=i+1, column=2).value = j.strip()

wb.save(filename = "C:\\Users\\shave\\Documents\\melodka_kveten.xlsx")
wb.close()
